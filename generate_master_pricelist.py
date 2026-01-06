import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os
import sys

def create_master_pricelist():
    print("--- Starting Excel Generator ---")
    
    # 1. PATH SETUP
    target_folder = r"C:\Users\GeorgeColes\OneDrive - Alder Technology\Documents\GCs Room Proposer"
    filename = "master_pricelist.xlsx"
    
    if os.path.exists(target_folder):
        full_path = os.path.join(target_folder, filename)
    else:
        print(f"Warning: Could not find '{target_folder}'. Saving to current folder.")
        full_path = "master_pricelist.xlsx"

    # 2. CREATE WORKBOOK
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pricelist"

    # 3. HEADERS
    headers = [
        "Max Distance",       # A
        "Tier Name",          # B
        "Cisco/VC Items",     # C (For Fit-Out, this is the Bar/VC Hardware)
        "Display Model",      # D
        "Display Price",      # E
        "Mount Model",        # F
        "Mount Price",        # G
        "Cables Desc",        # H
        "Cables Price",       # I
        "Service Price",      # J
        "MS Annual Price",    # K
        "Extra Items String"  # L (Format: Name|Cost|Qty; Name|Cost|Qty)
    ]

    # 4. DATA
    # Extras string format: "Item Name|Price|Qty ; Item Name|Price|Qty"
    xl_extras_98 = "QSC Core Nano|3500|1; Sennheiser Ceiling Mic|3576|2; QSC Ceiling Spk|765|6; Netgear Switch|1427|1; Wall Mount 82-98|100|1"
    xl_extras_86 = "QSC Core Nano|3500|1; Sennheiser Ceiling Mic|3576|2; QSC Ceiling Spk|765|6; Netgear Switch|1427|1; Wall Mount 82-98|100|1"

    data = [
        # [Dist, Name, VC_Item, Disp_Model, Disp_Price, Mnt_Model, Mnt_Price, Cab_Desc, Cab_Price, Svc_Price, MS_Price, Extras]
        
        # --- DATA#3 TIERS ---
        [3.0, "Small Meeting Space", "Cisco Room Bar", "LG 55UL3J-B", 1100, "Venturi VP-F80", 65, "HDMI & Patch Leads", 50, 3000, 1200, ""],
        [4.5, "Medium Meeting Space", "Cisco Room Bar, 1x Mic", "LG 65UL3J-B", 1500, "Venturi VP-F80", 65, "HDMI & Patch Leads", 50, 3000, 1200, ""],
        [5.5, "Large Meeting Space", "Cisco Room Bar Pro, 1x Mic", "LG 75UL3J-B", 2200, "Venturi VP-F80", 65, "HDMI & Fixings", 80, 3000, 1500, ""],
        [6.5, "Extra Large Space", "Cisco Room Bar Pro, 1x Mic", "LG 86UL3J-B", 3300, "VP-F100", 100, "HDMI & Fixings", 100, 3500, 1500, ""],
        [15.0, "Boardroom", "Cisco Kit EQ...", "LG 98UM5K", 7500, "VP-F100", 100, "HDMI & Spk Cable", 200, 4000, 3000, ""],

        # --- FIT-OUT TIERS ---
        [0.0, "Fit-Out 55", "Maxhub XBAR W70", "LG 55UL3J-B", 1100, "Venturi VP-F80", 65, "Custom Bundle", 300, 3000, 1200, ""],
        [0.0, "Fit-Out 65", "Maxhub XBAR W70", "LG 65UL3J-B", 1500, "Venturi VP-F80", 65, "Custom Bundle", 300, 3000, 1200, ""],
        [0.0, "Fit-Out 75", "Maxhub XBAR W70", "LG 75UL3J-B", 2200, "Venturi VP-F80", 65, "Custom Bundle", 300, 3000, 1200, ""],
        [0.0, "Fit-Out XL (98 Single)", "Maxhub XBAR W70", "LG 98UM5K", 9000, "Included in Extras", 0, "Custom Bundle", 1090, 9000, 1500, xl_extras_98],
        [0.0, "Fit-Out XL (86 Dual)", "Maxhub XBAR W70", "LG 86UL3J-B", 3300, "Included in Extras", 0, "Custom Bundle", 1090, 9500, 1500, xl_extras_86]
    ]

    ws.append(headers)
    for row in data:
        ws.append(row)

    # 5. STYLING
    header_fill = PatternFill(start_color="009A44", end_color="009A44", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    # 6. SAVE
    try:
        wb.save(full_path)
        print(f"SUCCESS! File saved at: {full_path}")
    except PermissionError:
        print("ERROR: The Excel file is open. Close it and try again.")
    except Exception as e:
        print(f"ERROR: {e}")

try:
    create_master_pricelist()
except Exception as e:
    print(f"\nCRITICAL ERROR: {e}")

input("\nPress Enter to close this window...")
